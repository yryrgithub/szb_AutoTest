import os
from openpyxl import load_workbook


class ExcelRead(object):

    def __init__(self, excel_path=None, sheet_name=None):
        if excel_path is None:
            current_path = os.path.dirname(os.path.dirname(__file__))
            self.excel_path = current_path + '/data/casedata.xlsx'
        else:
            self.excel_path = excel_path

        if sheet_name is None:
            self.sheet_name = 'Sheet1'
        else:
            self.sheet_name = sheet_name

        self.data = load_workbook(self.excel_path)
        self.sheet = self.data[self.sheet_name]

    def get_case(self):

        rows = self.sheet.rows
        row_num = self.sheet.max_row
        col_num = self.sheet.max_column
        if row_num <= 1:
            print("无数据")
        else:
            all_case = []
            for row in rows:
                case = []
                for i in range(col_num):
                    case.append(row[i].value)
                all_case.append(case)
            return all_case

    def get_case_account(self):
        # sheet = "Login"
        data = self.get_case()
        except_element_index = data[0].index('预期结果定位')
        except_index = data[0].index('预期结果')

        data_length = data.__len__()
        all_case = []
        for i in range(1, data_length):
            case = []
            for j in range(2,except_element_index):
                case.append(data[i][j])
            case.append(data[i][except_element_index])
            case.append(data[i][except_index])
            all_case.append(case)
        return all_case


if __name__ == '__main__':
    sheet = "Login"
    f = ExcelRead(sheet_name=sheet).get_case_account()
    print(f)
